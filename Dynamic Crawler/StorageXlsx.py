# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl import workbook


class Storage(object):

    def __init__(self):
        self.wb = openpyxl.load_workbook('mTime.xlsx')
        self.sheet = self.wb.get_sheet_by_name('Movies from mTime')
        self.datas = []

    def write_head(self):
        '''
        Create a headline for the table
        :return:
        '''
        heads = ['MovieId', 'MovieTitle', 'RatingFinal',
                 'ROtherFinal', 'RPictureFinal', 'RDirectorFinal',
                 'RStoryFinal', 'UserCount', 'AttitudeCount',
                 'TotalBoxOffice', 'TodayBoxOffice', 'Rank',
                 'Showdays', 'isRelease']
        for i, head in enumerate(heads):
            self.sheet.cell(row=1, column=i+1, value=head)

    def write_data(self, row, data):
        '''
        Write values into specific cells
        :param row:
        :param column:
        :param data: A tuple of values
        :return:
        '''
        for i, value in enumerate(data):
            self.sheet.cell(row=row, column=i+1, value=value)

    def write_end(self):
        '''
        Save the table at named file
        :return:
        '''
        self.wb.save('mTime.xlsx')




